import openpyxl
path = "/home/mrony/Salary_info.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook.active

row = sheet.max_row
column = sheet.max_column
print(row)
print(column)

for r in range(1,row+1):
    for c in range(1,column+1):
        print(sheet.cell(row=r, column=c).value)
    print()