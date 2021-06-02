import openpyxl

def getRowcount(file,sheetname):
   workbook = openpyxl.load_workbook(file)
   sheet= workbook.get_sheet_by_name(sheetname)
   return (sheet.max_row)

def getColumncount(file,sheetname):
     workbook = openpyxl.load_workbook(file)
     sheet= workbook.get_sheet_by_name(sheetname)
     return (sheet.max_column)

def readdata(file,sheetname,row,columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=row, column=columnnum) .value
def writedata(file,sheetname,row,columnnum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=row, column=columnnum).value = data
    return workbook.save(file)
