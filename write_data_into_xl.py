import openpyxl
path= "/home/mrony/data.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
row = sheet.max_row
col= sheet.max_column
for r in range(1,4):
    for c in range(1,6):
        sheet.cell(row=r, column=c).value = "Samrin"
workbook.save(path)
